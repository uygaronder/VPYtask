# Description: Transmits CSV containing vehicle information to the POST Call of the server & handles the response and generates an Excel-File
#   taking the input parameters into account.
# Usage: python client.py -k <keys> -c
# Keys: Keys to include as additional columns in the output Excel file
# Colored: Flag to indicate whether to color each row depending on the age of the hu field
# Example: python client.py -k 'make' 'model' 'color' -c
# Example to make it not colored (default): python client.py -k 'make' 'model' 'color'

import argparse
import csv
import datetime
import json
import requests
import openpyxl
from openpyxl.styles import PatternFill

# Hu colors
green_fill = PatternFill(start_color='007500', end_color='007500', fill_type='solid')
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
red_fill = PatternFill(start_color='b30000', end_color='b30000', fill_type='solid')

url = 'http://127.0.0.1:8000'

# Line Arguments
parser = argparse.ArgumentParser(description='Transmit CSV containing vehicle information to the POST Call of the server, handles the response and generates an Excel-File taking the input parameters into account.')
parser.add_argument('-k', '--keys', nargs='+', help='Keys to include as additional columns in the output Excel file')
parser.add_argument('-c', '--colored', action='store_true', default=True, help='Flag to indicate whether to color each row depending on the age of the hu field')
args = parser.parse_args()

column_headers = ['rnr']

with open('vehicles.csv', 'rb') as file:
    csv_data = file.read()
    payload = {'file': ('vehicles.csv', csv_data)}

# Send the POST request
response = requests.post(url+"/vehicles", files=payload)
data = json.loads(response.text)

if args.keys:
    for key in args.keys:
        column_headers.append(key)

workbook = openpyxl.Workbook()
worksheet = workbook.active

sorted_data = sorted(data["data"], key=lambda x: x.get('gruppe', ''))
column_headers = list(data["data"][0].keys())

# Write the column headers
for i, header in enumerate(column_headers):
    worksheet.cell(row=1, column=i + 1, value=header)

# Add the response fields as columns
for field in data["data"]:
    if field not in column_headers:
        column_headers.append(field)

# Write the data rows to the worksheet
for i, row in enumerate(sorted_data):
    row_index = i + 2
    try:
        for j, header in enumerate(column_headers):
            worksheet.cell(row=row_index, column=j + 1, value=row.get(header, ''))

            if args.colored and 'hu' in row:
                hu_date = datetime.datetime.strptime(row['hu'], '%Y-%m-%d')
                age_in_months = (datetime.datetime.now() - hu_date).days / 30

                if age_in_months <= 3:
                    worksheet.row_dimensions[row_index].fill = green_fill
                elif age_in_months <= 12:
                    worksheet.row_dimensions[row_index].fill = orange_fill
                else:
                    worksheet.row_dimensions[row_index].fill = red_fill
    except ValueError:
        continue
    

# Save the file
now = datetime.datetime.now()
formatted_timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
output_file_name = f'vehicles_{formatted_timestamp}.xlsx'
workbook.save(output_file_name)